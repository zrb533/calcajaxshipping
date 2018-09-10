# -*- coding=utf8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,colors
import re
from openpyxl.styles import numbers

source_data_path = '/Users/zhanglinquan/PycharmProjects/calcajaxshipping/source_data/dhl_express_remote_areas_201809041551451.xlsx'
target_data_path = '/Users/zhanglinquan/PycharmProjects/calcajaxshipping/target_data/target_data.xlsx'

eng_re = re.compile(r'^[A-Za-z,\s-]+$', re.S)

lw = load_workbook(source_data_path)

target_column = 1
wb = Workbook()
target_sheet = wb.active
target_sheet.title = "target_sheet"
ft_black = Font(color=colors.BLACK)
ft_red = Font(color=colors.RED)

for sheetnumber in range(1,192):
    print(sheetnumber)
    s_name = "Table " + str(sheetnumber)
    sheet = lw[s_name]

    max_row = sheet.max_row
    max_column = sheet.max_column

    for mc in range(1, max_column + 1):
        for mr in range(3, max_row + 1):
            cell = str(sheet.cell(mr,mc).value)
            cell_format = sheet.cell(mr,mc).number_format

            '''
            1. 先排除空行
            2. 再排除全英文（可以带空格和逗号），排除国家和州省数据
            3. 有区间的数据是按“空格-空格”分割，没有区间的数据直接输出（计数）
            4. 区间数据对分割后的左右数据进行去空格，逗号，减号操作
            5. 区间数据左右两边转成int类型，循环计数
            '''
            # 排除空格数据
            if cell == "None":
                pass
            else:
                cell = str(cell)
                res = re.findall(eng_re, cell)
                if len(res) > 0:
                    pass
                else:
                    if " - " in cell:
                        list_cell = cell.split(' - ')
                        cell_left = list_cell[0]
                        cell_right = list_cell[1]

                        # 如果左右两边的值中有空格或者横杠或逗号，去除
                        if "-" in cell_left or "-" in cell_right:
                            cell_left = cell_left.replace("-","")
                            cell_right = cell_right.replace("-", "")
                        if " " in cell_left or " " in cell_right:
                            cell_left = cell_left.replace(" ","")
                            cell_right = cell_right.replace(" ", "")
                        if "," in cell_left or "," in cell_right:
                            cell_left = cell_left.replace(",","")
                            cell_right = cell_right.replace(",", "")

                        int_left = int(cell_left)
                        int_right = int(cell_right)
                        for a in range(int_left, int_right + 1):
                            a = str(a)
                            # target_sheet['A'+str(target_column)] = a
                            target_column = target_column + 1
                    else:
                        # 如果左右两边的值中有空格或者横杠或逗号，去除
                        if "-" in cell:
                            cell = cell.replace("-", "")
                        if " " in cell:
                            cell = cell.replace(" ", "")
                        if "," in cell:
                            cell = cell.replace(",", "")
                        # target_sheet['A' + str(target_column)] = cell
                        target_column = target_column + 1

    wb.save(target_data_path)
print(target_column)